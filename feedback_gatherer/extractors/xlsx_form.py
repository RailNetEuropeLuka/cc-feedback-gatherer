"""Extract the MS Form export: one .xlsx whose columns are keyed by guideline
section. Each data row is one respondent; each non-empty section cell is one item.

Unlike the other extractors, this returns ONE ParsedSource per row (a single file
holds many respondents).
"""
from __future__ import annotations

import io
import re

import openpyxl

from .base import ParsedSource, RawItem


def _is_gibberish(text: str, min_len: int) -> bool:
    """Heuristic test-row detector: short token, no spaces, no vowels."""
    t = (text or "").strip()
    if len(t) < min_len:
        return False
    if " " in t:
        return False
    return not re.search(r"[aeiouAEIOU]", t)


def looks_like_form_export(headers: list) -> bool:
    joined = " ".join(str(h or "").lower() for h in headers)
    return "company" in joined and ("incentives" in joined or "goal of commercial" in joined)


def extract(data: bytes, filename: str, cfg: dict) -> list[ParsedSource]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.worksheets[0]
    headers = [c.value for c in ws[1]]

    mf = cfg.get("msform_export", {})
    # Prefer configured columns; fall back to header auto-detection.
    if looks_like_form_export(headers):
        company_col = mf.get("company_col", _find(headers, "company"))
        email_col = mf.get("email_col", _find(headers, "e-mail:"))
        fn_col = mf.get("firstname_col", _find(headers, "first name"))
        ln_col = mf.get("lastname_col", _find(headers, "last name"))
        date_col = mf.get("date_col", _find(headers, "fertigstellung"))
        sec_lo = mf.get("first_section_col", 10)
        sec_hi = mf.get("last_section_col", 21)
        section_cols = list(range(sec_lo, sec_hi + 1))
    else:
        # generic spreadsheet: treat every column whose header starts with a
        # section number as an answer column.
        section_cols = [i for i, h in enumerate(headers)
                        if re.match(r"\s*[1-3]\.\d", str(h or ""))]
        company_col = _find(headers, "company")
        email_col = _find(headers, "mail")
        fn_col = ln_col = date_col = None

    tf = cfg.get("test_row_filters", {})
    internal = set(d.lower() for d in tf.get("internal_domains", []))
    anon = set(v.lower() for v in tf.get("anonymous_email_values", []))
    test_companies = set(v.lower() for v in tf.get("test_company_values", []))
    test_names = set(v.lower() for v in tf.get("test_name_values", []))
    gib_len = tf.get("gibberish_min_len", 4)

    sources: list[ParsedSource] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        company = _cell(row, company_col)
        email = _cell(row, email_col)
        first = _cell(row, fn_col)
        last = _cell(row, ln_col)
        rep = (f"{first} {last}".strip() or None) if (first or last) else None
        date = _cell(row, date_col)

        ps = ParsedSource(source_format="xlsx",
                           company_hint=company or None,
                           company_authoritative=True,  # the Company column is reliable
                           email_hint=email or None,
                           date_hint=_iso(date))
        if rep:
            ps.notes.append(f"rep:{rep}")

        # is this a test row?
        dom = email.split("@")[-1].lower() if "@" in email else ""
        answers = [(_header(headers, c), _cell(row, c)) for c in section_cols]
        answers = [(h, v) for h, v in answers if v.strip()]
        test = (
            email.lower() in anon
            or dom in internal
            or company.lower() in test_companies
            or first.lower() in test_names
            or (bool(answers) and all(_is_gibberish(v, gib_len) for _, v in answers))
        )
        if test:
            ps.notes.append("Skipped: internal/test submission.")
            ps.full_text = "\n".join(f"{h}: {v}" for h, v in answers)
            ps.is_endorsement = False
            ps.items = []
            ps.notes.append("__TEST__")
            sources.append(ps)
            continue

        for header, value in answers:
            ps.items.append(RawItem(section_raw=str(header), considerations=value,
                                    raw_text=f"{header}: {value}"))
        ps.full_text = "\n\n".join(f"### {h}\n{v}" for h, v in answers)
        if not ps.items:
            ps.notes.append("No section answers filled in.")
        sources.append(ps)
    return sources


# --- small helpers ----------------------------------------------------------
def _find(headers: list, needle: str) -> int | None:
    needle = needle.lower()
    for i, h in enumerate(headers):
        if h and needle in str(h).lower():
            return i
    return None


def _cell(row, idx) -> str:
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


def _header(headers, idx) -> str:
    return str(headers[idx]) if idx < len(headers) and headers[idx] else ""


def _iso(v) -> str | None:
    if v is None:
        return None
    s = str(v)
    m = re.search(r"\d{4}-\d{2}-\d{2}", s)
    return m.group(0) if m else None
