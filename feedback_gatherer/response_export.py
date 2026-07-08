"""Task Force response workbook (stage 2.2).

Builds the Excel file the team downloads from the dashboard to draft their
answers to the consultation feedback: one sheet per Guidelines chapter, one
ANSWER ROW per recurring point (with the grouped comments listed underneath)
plus one per unique point - so a point raised by four organisations is answered
once, not four times.

Headless: takes the DataFrames the dashboard already computes, returns bytes.
Never writes to disk - the workbook only exists in the download.
"""
from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from analysis import SECTION_ORDER

STATUS_VALUES = ["Accepted", "Partially accepted", "Rejected",
                 "Clarification provided", "Out of scope"]

COLUMNS = [  # (header, width)
    ("Unit", 10), ("Kind", 26), ("Organisation(s)", 30), ("Feedback", 80),
    ("TF Response", 60), ("Status", 20), ("Assignee", 14), ("Notes", 30),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_PARENT_FILL = PatternFill("solid", fgColor="EAF1FA")   # answer rows stand out
_MEMBER_FONT = Font(color="52514E")
_GREY_FILL = PatternFill("solid", fgColor="F0EFEC")     # non-answer cells

_WRAP = Alignment(wrap_text=True, vertical="top")


def _sheet_header(ws):
    ws.append([c for c, _ in COLUMNS])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _status_validation(ws) -> DataValidation:
    dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_VALUES) + '"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    return dv


def _append(ws, values, *, parent: bool, member: bool = False):
    ws.append(values)
    row = ws.max_row
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row, col)
        cell.alignment = _WRAP
        if parent and col <= 4:
            cell.fill = _PARENT_FILL
        if member:
            cell.font = _MEMBER_FONT
            if col >= 5:
                cell.fill = _GREY_FILL     # answer lives on the parent row
    return row


def build_response_workbook(items_df: pd.DataFrame, clustered: pd.DataFrame,
                            themes: pd.DataFrame) -> bytes:
    wb = Workbook()

    # ------------------------------------------------------------- How to use
    ws = wb.active
    ws.title = "How to use"
    intro = [
        ["Task Force response workbook"],
        [f"Generated {datetime.now():%d %b %Y %H:%M} from {len(items_df)} MS Form "
         f"feedback items."],
        [""],
        ["One sheet per Guidelines chapter. Each sheet contains ANSWER ROWS (blue):"],
        ["  - 'Recurring point - N organisations': several organisations made this "
         "point; the grouped comments are listed underneath (grey). Answer ONCE on "
         "the blue row."],
        ["  - 'Single point': raised by one organisation only."],
        [""],
        ["Fill in: TF Response, Status (dropdown), Assignee, Notes."],
        ["Status values: " + " / ".join(STATUS_VALUES) + "."],
        [""],
        ["The grouping is machine-suggested. If one grouped comment needs a "
         "different answer, write it in that comment's Notes cell or split the "
         "response."],
        ["The 'All items' sheet lists every comment with the unit it belongs to."],
    ]
    for line in intro:
        ws.append(line)
    ws.column_dimensions["A"].width = 110
    for r in range(1, ws.max_row + 1):
        ws.cell(r, 1).alignment = _WRAP
    ws["A1"].font = Font(bold=True, size=14)

    # --------------------------------------------------------- chapter sheets
    item_unit: dict[str, str] = {}          # item_id -> unit_id (for All items)
    sections = [s for s in SECTION_ORDER if s in set(items_df["section_ref"])]
    for sec in sections:
        title = items_df.loc[items_df["section_ref"] == sec, "section_title"].iloc[0]
        ws = wb.create_sheet(sec)
        _sheet_header(ws)
        dv = _status_validation(ws)

        sec_themes = themes[themes["cluster_id"].str.startswith(f"{sec}/")] \
            if not themes.empty else themes
        n_p = 0
        for _, t in sec_themes.sort_values("n_respondents", ascending=False).iterrows():
            n_p += 1
            unit = f"{sec}-P{n_p}"
            medoid = clustered[clustered["item_id"] == t["medoid_item_id"]].iloc[0]
            row = _append(ws, [unit, f"Recurring point - {t['n_respondents']} organisations",
                               ", ".join(t["respondents"]),
                               medoid["considerations"], "", "", "", ""], parent=True)
            dv.add(ws.cell(row, 6))
            for i in t["member_idx"]:
                m = clustered.iloc[i]
                item_unit[m["item_id"]] = unit
                _append(ws, ["", "  ↳ part of the point above", m["company"],
                             m["considerations"], "", "", "", ""],
                        parent=False, member=True)

        singles = clustered[(clustered["section_ref"] == sec)
                            & (clustered["cluster_id"] == "-")]
        n_u = 0
        for _, s in singles.iterrows():
            n_u += 1
            unit = f"{sec}-U{n_u}"
            item_unit[s["item_id"]] = unit
            row = _append(ws, [unit, "Single point", s["company"],
                               s["considerations"], "", "", "", ""], parent=True)
            dv.add(ws.cell(row, 6))

        ws.append([])
        ws.append(["", f"Chapter {sec} - {title}: {n_p} recurring point(s), "
                       f"{n_u} single point(s)."])

    # -------------------------------------------------------------- All items
    ws = wb.create_sheet("All items")
    ref_cols = [("item_id", 24), ("unit", 10), ("chapter", 9),
                ("organisation", 28), ("type", 13), ("feedback", 90)]
    ws.append([c for c, _ in ref_cols])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, (_, width) in enumerate(ref_cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    for _, r in clustered.iterrows():
        ws.append([r["item_id"], item_unit.get(r["item_id"], ""), r["section_ref"],
                   r["company"], r["classification"], r["considerations"]])
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(ref_cols) + 1):
            ws.cell(row, col).alignment = _WRAP

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
