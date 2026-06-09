"""Serialise gathered feedback to JSON + Excel. Used by the CLI (write to disk)
and by the web app (build in-memory bytes for download)."""
from __future__ import annotations

import io
import json

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from schema import FeedbackItem, Respondent, build_envelope

ITEM_COLUMNS = [
    ("item_id", 18), ("respondent_id", 18), ("company", 26), ("classification", 13),
    ("section_ref", 11), ("section_title", 30), ("section_raw", 18),
    ("considerations", 70), ("proposal", 50), ("channel", 10),
    ("source_file", 34), ("source_format", 8), ("date", 12),
    ("fte_alignment", 13), ("endorses", 12), ("extraction_confidence", 12),
    ("needs_review", 12), ("review_note", 28),
]
RESP_COLUMNS = [
    ("respondent_id", 18), ("company", 28), ("classification", 13),
    ("representative", 22), ("email", 28), ("country", 12),
    ("channels", 18), ("source_files", 50), ("fte_alignment", 13), ("n_items", 9),
]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def to_json_bytes(items: list[FeedbackItem], respondents: list[Respondent],
                  meta: dict | None = None) -> bytes:
    env = build_envelope(items, respondents, meta)
    return json.dumps(env, ensure_ascii=False, indent=2).encode("utf-8")


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def _write_sheet(ws, columns, rows: list[dict]):
    ws.append([c for c, _ in columns])
    for row in rows:
        vals = []
        for key, _ in columns:
            v = row.get(key)
            if isinstance(v, list):
                v = ", ".join(str(x) for x in v)
            vals.append(v)
        ws.append(vals)
    for i, (_, width) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    _style_header(ws)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(columns) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")


def to_excel_bytes(items: list[FeedbackItem], respondents: list[Respondent]) -> bytes:
    wb = Workbook()
    ws_items = wb.active
    ws_items.title = "Items"
    _write_sheet(ws_items, ITEM_COLUMNS, [i.to_dict() for i in items])

    ws_resp = wb.create_sheet("Respondents")
    _write_sheet(ws_resp, RESP_COLUMNS, [r.to_dict() for r in respondents])

    ws_src = wb.create_sheet("Summary")
    _write_summary(ws_src, items, respondents)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary(ws, items: list[FeedbackItem], respondents: list[Respondent]):
    ws.append(["Metric", "Value"])
    ws.append(["Total feedback items", len(items)])
    ws.append(["Respondents with feedback", len(respondents)])
    ws.append([])
    ws.append(["Items per guideline section", ""])
    by_sec: dict[str, int] = {}
    for it in items:
        by_sec[it.section_ref] = by_sec.get(it.section_ref, 0) + 1
    for sec in sorted(by_sec):
        ws.append([sec, by_sec[sec]])
    ws.append([])
    ws.append(["Items per classification", ""])
    by_cls: dict[str, int] = {}
    for it in items:
        by_cls[it.classification] = by_cls.get(it.classification, 0) + 1
    for cls in sorted(by_cls):
        ws.append([cls, by_cls[cls]])
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    _style_header(ws)
